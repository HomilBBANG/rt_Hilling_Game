#!/usr/bin/env python3
"""게임 밸런스: 엑셀(.xlsx) -> Godot용 JSON 변환기.

data/balance.xlsx 를 읽어 data/balance.json (UTF-8) 을 생성합니다.
엑셀에서 수치를 수정한 뒤 이 스크립트를 실행하면 게임 값이 갱신됩니다.

    사용법:  python tools/convert_balance.py

필요 패키지:  pip install openpyxl

엑셀 시트 첫 행은 헤더여야 하며, 다음 열을 사용합니다(대소문자 무시):
    key   : 수치 이름(영문) — 필수
    value : 값(숫자) — 필수
    note  : 설명(게임에는 미사용)
"""
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("[에러] openpyxl 이 필요합니다.  pip install openpyxl")
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "data", "balance.xlsx")
OUT = os.path.join(ROOT, "data", "balance.json")


def coerce(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else v
    s = str(v).strip()
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return s


def main() -> int:
    if not os.path.exists(XLSX):
        print(f"[에러] 엑셀 파일 없음: {XLSX}")
        return 1

    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("[에러] 시트가 비어 있습니다.")
        return 1

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "key" not in header or "value" not in header:
        print(f"[에러] 헤더에 'key' 와 'value' 가 필요합니다. 현재 헤더: {header}")
        return 1
    ki, vi = header.index("key"), header.index("value")

    values = {}
    for r in rows[1:]:
        if r is None:
            continue
        key = r[ki] if ki < len(r) else None
        val = r[vi] if vi < len(r) else None
        if key is None or str(key).strip() == "" or val is None:
            continue
        values[str(key).strip()] = coerce(val)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(values, f, ensure_ascii=False, indent=2)
    print(f"[완료] {len(values)}개 값 → {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
