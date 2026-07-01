#!/usr/bin/env python3
"""NPC 해금 표: 엑셀(.xlsx) -> Godot용 JSON 변환기.

data/npc_unlocks.xlsx 를 읽어 data/npc_unlocks.json (UTF-8) 을 생성합니다.
엑셀에서 임계값/NPC를 수정한 뒤 이 스크립트를 실행하면 게임 데이터가 갱신됩니다.

    사용법:  python tools/convert_npc_unlocks.py

필요 패키지:  pip install openpyxl

엑셀 시트 첫 행은 헤더여야 하며, 다음 열을 사용합니다(대소문자 무시):
    threshold      : 누적 만족도 임계값(정수) — 필수
    npc_id         : NPC 식별자(영문) — 필수
    display_name   : 표시 이름(한글 가능)
    role           : unique | placeable
    unlocks        : 해금 콘텐츠 id
    note           : 메모(게임에는 미사용)
"""
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("[에러] openpyxl 이 필요합니다.  pip install openpyxl")
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "data", "npc_unlocks.xlsx")
OUT = os.path.join(ROOT, "data", "npc_unlocks.json")

OPTIONAL_COLS = ["display_name", "role", "unlocks", "note"]


def main() -> int:
    if not os.path.exists(XLSX):
        print(f"[에러] 엑셀 파일 없음: {XLSX}")
        return 1

    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("[에러] 시트가 비어 있습니다.")
        return 1

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "threshold" not in header or "npc_id" not in header:
        print(f"[에러] 헤더에 'threshold' 와 'npc_id' 가 필요합니다. 현재 헤더: {header}")
        return 1
    idx = {name: header.index(name) for name in header if name}

    entries = []
    for r in rows[1:]:
        if r is None:
            continue
        npc_id = r[idx["npc_id"]] if idx.get("npc_id") is not None else None
        thr = r[idx["threshold"]] if idx.get("threshold") is not None else None
        if npc_id is None or str(npc_id).strip() == "" or thr is None:
            continue
        entry = {"threshold": int(thr), "npc_id": str(npc_id).strip()}
        for col in OPTIONAL_COLS:
            if col in idx and r[idx[col]] is not None:
                entry[col] = str(r[idx[col]]).strip()
        entries.append(entry)

    entries.sort(key=lambda e: e["threshold"])

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)
    print(f"[완료] {len(entries)}개 항목 → {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
